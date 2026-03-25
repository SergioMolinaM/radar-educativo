"""Exports router - PDF and Excel report generation."""
import io
import logging
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from api.routers.auth import get_current_user
from api.routers.dashboard import get_dashboard_summary, get_semaforos

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/excel")
def export_excel(current_user: dict = Depends(get_current_user)):
    """Export dashboard data as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    summary = get_dashboard_summary(current_user)
    semaforos = get_semaforos(current_user)

    wb = Workbook()

    # Sheet 1: Resumen KPIs
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_properties.tabColor = "3b82f6"

    header_font = Font(bold=True, size=14, color="1e40af")
    subheader_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws1["A1"] = "Radar Educativo - Reporte SLEP"
    ws1["A1"].font = header_font
    ws1["A2"] = f"SLEP: {current_user['slep_id'].upper()} | Fecha: {datetime.now().strftime('%d/%m/%Y')}"
    ws1["A2"].font = Font(color="666666")

    row = 4
    kpis = summary.get("kpis", {})
    kpi_items = [
        ("Total establecimientos", kpis.get("total_establecimientos", 0)),
        ("Matrícula total", kpis.get("matricula_total", 0)),
        ("Asistencia promedio (%)", kpis.get("asistencia_promedio", 0)),
        ("Ejecución presupuestaria (%)", kpis.get("ejecucion_presupuestaria", 0)),
        ("Alertas rojas", kpis.get("alertas_rojas", 0)),
        ("Alertas naranjas", kpis.get("alertas_naranjas", 0)),
        ("Alertas verdes", kpis.get("alertas_verdes", 0)),
    ]
    for label, value in kpi_items:
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=value)
        row += 1

    # Sheet 2: Establecimientos con semáforo
    ws2 = wb.create_sheet("Establecimientos")
    headers = ["RBD", "Nombre", "Matrícula", "Asistencia %", "Ejecución %", "Semáforo", "Alertas"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = subheader_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    semaforo_fills = {
        "rojo": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "naranja": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "verde": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    }

    for i, est in enumerate(semaforos.get("establecimientos", []), 2):
        ws2.cell(row=i, column=1, value=est.get("rbd")).border = thin_border
        ws2.cell(row=i, column=2, value=est.get("nombre")).border = thin_border
        ws2.cell(row=i, column=3, value=est.get("matricula")).border = thin_border
        ws2.cell(row=i, column=4, value=est.get("asistencia")).border = thin_border
        ws2.cell(row=i, column=5, value=est.get("ejecucion")).border = thin_border

        sem_cell = ws2.cell(row=i, column=6, value=est.get("semaforo", "").upper())
        sem_cell.fill = semaforo_fills.get(est.get("semaforo"), PatternFill())
        sem_cell.font = Font(bold=True)
        sem_cell.alignment = Alignment(horizontal="center")
        sem_cell.border = thin_border

        ws2.cell(row=i, column=7, value="; ".join(est.get("alertas", []))).border = thin_border

    # Auto-width columns
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"radar_educativo_{current_user['slep_id']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
